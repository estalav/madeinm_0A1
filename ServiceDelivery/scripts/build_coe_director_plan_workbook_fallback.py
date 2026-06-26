from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


ROOT = Path.cwd()
OUTPUT_DIR = ROOT / "outputs" / "coe-director-plan"
OUTPUT_PATH = OUTPUT_DIR / "TesseraLabs_IO_CoE_Director_Implementation_Plan_July_2026.xlsx"

ROADMAP = [
    [
        "July 2026",
        "Mobilize",
        "Establish the CoE charter, sponsors, and first service priorities.",
        "Confirm CoE mission and scope; review product positioning and sales narrative; select Wave 1 services; baseline existing assets; define governance cadence.",
        "CoE charter; stakeholder map; 6-month roadmap; service prioritization decision; initial meeting cadence.",
        "Executive alignment achieved; scope approved; workstreams launched.",
    ],
    [
        "August 2026",
        "Design",
        "Design the delivery framework, governance model, and KPI/SLA baseline.",
        "Define service catalog and tiers; define delivery lifecycle; create RACI; define reporting cadence and escalation paths; create template inventory.",
        "Service catalog v1; delivery methodology v1; RACI model; governance pack; KPI and SLA framework.",
        "Framework approved for pilot use; service definitions usable by sales and delivery.",
    ],
    [
        "September 2026",
        "Build",
        "Create the asset library and pilot playbooks needed to run delivery consistently.",
        "Build onboarding and discovery templates; build scope templates; create RAID, backlog, status, and readiness trackers; finalize Wave 1 playbooks; define staffing and handoff model.",
        "Delivery asset library; pilot playbooks; service review templates; staffing and handoff model.",
        "Core templates available; pilot teams can execute with a standard toolkit.",
    ],
    [
        "October 2026",
        "Pilot",
        "Validate the CoE through early customer work or structured dry runs.",
        "Run one or two pilot engagements; test discovery, planning, wave management, and reporting; validate governance flow; capture operating-model gaps.",
        "Pilot execution report; issue log; improvement backlog; updated methodology and templates.",
        "Pilot work executed with measurable control; operating gaps identified and prioritized.",
    ],
    [
        "November 2026",
        "Stabilize",
        "Harden the model after pilot feedback and prepare for repeat use.",
        "Refine playbooks and templates; align product, sales, and delivery handoffs; define quality checkpoints; implement monthly service review structure; finalize Wave 2 assumptions.",
        "CoE operating model v2; quality assurance framework; handoff governance model; Wave 2 design brief.",
        "CoE stable enough for repeat use; cross-functional handoffs clarified.",
    ],
    [
        "December 2026",
        "Scale Readiness",
        "Prepare the CoE to scale in 2027 using evidence from the first execution cycles.",
        "Define the 2027 roadmap; define hiring priorities and partner needs; baseline KPIs; finalize dashboards; publish reusable accelerators and lessons learned.",
        "2027 scale roadmap; hiring and capability plan; KPI baseline report; final asset repository structure.",
        "Repeatable operating backbone in place; leadership has a scale plan and performance baseline.",
    ],
]

WORKSTREAMS = [
    ["Service Catalog and Packaging", "Define services, scope boundaries, service tiers, and value propositions."],
    ["Delivery Methodology", "Define lifecycle, phase gates, quality checkpoints, and readiness criteria."],
    ["Operating Model and Governance", "Define decision rights, cadence, escalation, RAID process, and reporting."],
    ["Delivery Assets and Accelerators", "Create onboarding templates, scope tools, dashboards, and trackers."],
    ["KPI, SLA, and Quality Management", "Create the minimum viable performance and quality system for delivery."],
    ["Pilot Delivery Enablement", "Prepare pilot playbooks, staffing, hypercare, and lessons learned flow."],
    ["CoE Team and Capability Build", "Define the minimum viable team, capability matrix, and training plan."],
]

FIRST_30_DAYS = [
    ["Week 1", "Align with founders on mandate, success criteria, product messaging, and current pipeline."],
    ["Week 2", "Assess delivery maturity, identify gaps, define Wave 1 services, map stakeholders."],
    ["Week 3", "Draft CoE charter, service catalog, governance cadence, and pilot role needs."],
    ["Week 4", "Review CoE design with leadership, confirm pilot priorities, publish roadmap, launch workstreams."],
]

KPIS = [
    "Percentage of priority services with approved playbooks",
    "Percentage of delivery templates completed",
    "Average time to mobilize a new engagement",
    "Percentage of engagements using the standard methodology",
    "Scope change rate after kickoff",
    "Milestone adherence rate",
    "Defect leakage rate from remediation to validation",
    "Customer reporting timeliness",
    "SLA compliance by service type",
    "Lessons learned closed within target window",
]

RISKS = [
    ["Product and service are not translated clearly enough", "Map platform capabilities to sellable services, delivery activities, and measurable outcomes."],
    ["Sales closes work before delivery is standardized", "Introduce scope guardrails and delivery approval checkpoints before deal commitment."],
    ["Too many services are launched at once", "Focus on Wave 1 services first and delay broader standardization until pilot evidence exists."],
    ["Early customer work becomes fully custom", "Use Fit-to-Standard principles, phase gates, and governance approval for exceptions."],
    ["Delivery metrics are not defined early", "Launch with a minimum viable KPI and SLA pack before the first engagement starts."],
]


THIN = Side(style="thin", color="C7D0DA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(cell, fill_color):
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_body(cell, wrap=True):
    cell.alignment = Alignment(wrap_text=wrap, vertical="top")
    cell.border = BORDER


def set_widths(ws, widths):
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    overview = wb.active
    overview.title = "Overview"
    roadmap = wb.create_sheet("Roadmap")
    workstreams = wb.create_sheet("Workstreams")
    first30 = wb.create_sheet("First 30 Days")
    kpi_risk = wb.create_sheet("KPIs and Risks")

    overview.merge_cells("A1:F1")
    overview["A1"] = "TesseraLabs.IO Director of CoE Implementation Plan"
    overview["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    overview["A1"].fill = PatternFill("solid", fgColor="0B3558")
    overview["A1"].alignment = Alignment(horizontal="center")

    overview.merge_cells("A2:F2")
    overview["A2"] = "Start date: July 2026 | Focus: service delivery operating model for sold services"
    overview["A2"].fill = PatternFill("solid", fgColor="DDECF8")
    overview["A2"].font = Font(italic=True, color="23405A")
    overview["A2"].alignment = Alignment(horizontal="center")

    overview_rows = [
        ["Category", "Summary", "", "Key Deliverables", "By December 2026", ""],
        ["Mission", "Industrialize service delivery and convert platform capabilities into repeatable customer services.", "", "1", "Standard service catalog", ""],
        ["Wave 1", "ERP Migration Acceleration; Code Remediation and Optimization", "", "2", "Delivery methodology and phase gates", ""],
        ["Wave 2", "Data Quality and Harmonization; Cross-System Orchestration; Continuity and Operations Support", "", "3", "Governance, RACI, and escalation model", ""],
        ["Primary Outcome", "Repeatable, governed, measurable service delivery by December 2026", "", "4", "Pilot-ready playbooks for 2 priority services", ""],
        ["Core Role", "Director of Center of Excellence", "", "5", "KPI and SLA framework", ""],
        ["", "", "", "6", "Reusable delivery assets and dashboards", ""],
        ["", "", "", "7", "2027 scale roadmap", ""],
    ]

    for row_idx, row in enumerate(overview_rows, start=4):
        for col_idx, value in enumerate(row, start=1):
            cell = overview.cell(row=row_idx, column=col_idx, value=value)
            style_body(cell)
    style_header(overview["A4"], "0F766E")
    style_header(overview["B4"], "0F766E")
    style_header(overview["D4"], "2563EB")
    style_header(overview["E4"], "2563EB")
    set_widths(overview, {"A": 20, "B": 60, "C": 4, "D": 14, "E": 42, "F": 4})

    roadmap_headers = ["Month", "Phase", "Objective", "Key Activities", "Deliverables", "Success Criteria"]
    roadmap.append(roadmap_headers)
    for item in ROADMAP:
        roadmap.append(item)
    for cell in roadmap[1]:
        style_header(cell, "0B3558")
    for row in roadmap.iter_rows(min_row=2):
        for cell in row:
            style_body(cell)
    set_widths(roadmap, {"A": 16, "B": 18, "C": 42, "D": 65, "E": 55, "F": 42})

    workstreams.append(["Workstream", "Purpose"])
    for item in WORKSTREAMS:
        workstreams.append(item)
    for cell in workstreams[1]:
        style_header(cell, "0F766E")
    for row in workstreams.iter_rows(min_row=2):
        for cell in row:
            style_body(cell)
    set_widths(workstreams, {"A": 36, "B": 70})

    first30.append(["Week", "Focus"])
    for item in FIRST_30_DAYS:
        first30.append(item)
    for cell in first30[1]:
        style_header(cell, "7C3AED")
    for row in first30.iter_rows(min_row=2):
        for cell in row:
            style_body(cell)
    set_widths(first30, {"A": 12, "B": 82})

    kpi_risk.append(["KPI", "Risk", "Mitigation"])
    for index, kpi in enumerate(KPIS):
        risk = RISKS[index] if index < len(RISKS) else ["", ""]
        kpi_risk.append([kpi, risk[0], risk[1]])
    for cell in kpi_risk[1]:
        style_header(cell, "B45309")
    for row in kpi_risk.iter_rows(min_row=2):
        for cell in row:
            style_body(cell)
    set_widths(kpi_risk, {"A": 44, "B": 42, "C": 62})

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

    wb.save(OUTPUT_PATH)
    print(f"Created {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
